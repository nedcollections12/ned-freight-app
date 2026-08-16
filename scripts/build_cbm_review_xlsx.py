"""
Turn cbm_comparison.csv into a review-friendly Excel workbook for manual CBM cleanup.

Adds a plain-English "Review flag" per row and colour-codes issues so staff can
fix the Cin7 source values before the automated sync is switched on.

Usage:
    python scripts/build_cbm_review_xlsx.py [--in cbm_comparison.csv] [--out cbm_comparison.xlsx]
"""

import csv
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent.parent

# Largest genuine NED item (Montana/Dyne sofa) is ~2.5 m3; anything above this in
# Cin7 is almost certainly a bad default (e.g. 31.0) or a misplaced decimal.
MAX_PLAUSIBLE_CBM = 4.0

FILLS = {
    "bad":     PatternFill("solid", fgColor="F4B0B0"),  # red  — implausible Cin7 value
    "differs": PatternFill("solid", fgColor="FFE49C"),  # amber — real difference to verify
    "cin7zero":PatternFill("solid", fgColor="E8E8E8"),  # grey — Shopify has value, Cin7 blank
    "bothzero":PatternFill("solid", fgColor="F7D6C4"),  # peach — no CBM anywhere
    "match":   PatternFill("solid", fgColor="CDE9CB"),  # green — aligned
}
HEADER_FILL = PatternFill("solid", fgColor="1F2A44")


def _f(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def review_flag(row):
    status = row["status"]
    c = _f(row["cin7_cbm"])
    if status == "DIFFERS" and c is not None and c > MAX_PLAUSIBLE_CBM:
        return "FIX IN CIN7: CBM implausible (>4 m3) — looks like a default/kg or misplaced decimal", "bad"
    if status == "DIFFERS":
        return "VERIFY: Cin7 differs from Shopify — confirm correct value, fix in Cin7", "differs"
    if status == "CIN7_ZERO":
        return "Cin7 has no CBM — enter it in Cin7 (Shopify currently carries the value)", "cin7zero"
    if status == "BOTH_ZERO":
        return "No CBM anywhere — enter it in Cin7", "bothzero"
    if status == "MISSING_IN_SHOPIFY":
        return "Cin7 SKU with no active Shopify variant — no action (wholesale/discontinued?)", None
    if status == "MISSING_IN_CIN7":
        return "Shopify variant with no Cin7 SKU match — check SKU alignment", None
    return "OK — matches", "match"


def main():
    in_path = ROOT / "cbm_comparison.csv"
    out_path = ROOT / "cbm_comparison.xlsx"
    if "--in" in sys.argv:
        in_path = Path(sys.argv[sys.argv.index("--in") + 1])
    if "--out" in sys.argv:
        out_path = Path(sys.argv[sys.argv.index("--out") + 1])

    rows = list(csv.DictReader(in_path.open()))

    wb = Workbook()

    # ── Sheet 1: full comparison ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "CBM Comparison"
    headers = ["SKU", "Product", "Option", "Cin7 CBM", "Shopify CBM",
               "Diff (Cin7-Shopify)", "Status", "Review flag"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")

    counts = {}
    for r in rows:
        counts[r["status"]] = counts.get(r["status"], 0) + 1
        flag, fillkey = review_flag(r)
        ws.append([
            r["sku"], r["product"], r["option"],
            _f(r["cin7_cbm"]) if r["cin7_cbm"] != "" else "",
            _f(r["shopify_cbm"]) if r["shopify_cbm"] != "" else "",
            _f(r["diff_cin7_minus_shopify"]) if r["diff_cin7_minus_shopify"] != "" else "",
            r["status"], flag,
        ])
        if fillkey:
            for c in range(1, len(headers) + 1):
                ws.cell(ws.max_row, c).fill = FILLS[fillkey]

    widths = [16, 28, 22, 11, 12, 16, 18, 62]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    # ── Sheet 2: summary + legend ─────────────────────────────────────────────
    s = wb.create_sheet("Summary")
    s.append(["CBM comparison — Cin7 (source) vs Shopify (freight app)"])
    s["A1"].font = Font(bold=True, size=13)
    s.append([])
    s.append(["Status", "Count", "What it means / action"])
    for c in range(1, 4):
        s.cell(3, c).font = Font(bold=True, color="FFFFFF")
        s.cell(3, c).fill = HEADER_FILL
    meaning = {
        "DIFFERS": "Cin7 value differs from Shopify — review & fix in Cin7 (red rows = implausible)",
        "CIN7_ZERO": "Shopify has a CBM but Cin7 is blank — enter the CBM in Cin7",
        "BOTH_ZERO": "No CBM anywhere — enter it in Cin7",
        "MISSING_IN_SHOPIFY": "Cin7 SKU with no active Shopify variant — usually no action",
        "MISSING_IN_CIN7": "Shopify variant with no Cin7 match — check SKU alignment",
        "MATCH": "Already aligned — no action",
    }
    order = ["DIFFERS", "CIN7_ZERO", "BOTH_ZERO", "MISSING_IN_SHOPIFY", "MISSING_IN_CIN7", "MATCH"]
    for st in order:
        s.append([st, counts.get(st, 0), meaning[st]])
    s.append([])
    s.append([f"Plausibility ceiling used for flags: {MAX_PLAUSIBLE_CBM} m3 "
              "(bigger than the largest real NED item ~2.5 m3)."])
    s.column_dimensions["A"].width = 20
    s.column_dimensions["B"].width = 8
    s.column_dimensions["C"].width = 70

    wb.save(out_path)
    print(f"Wrote {out_path}  ({ws.max_row - 1} rows)")
    for st in order:
        print(f"  {st:20} {counts.get(st, 0)}")


if __name__ == "__main__":
    main()
