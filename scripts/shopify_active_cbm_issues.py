"""
From cbm_comparison.csv, produce the list of ACTIVE Shopify variants whose CBM is
missing or disagrees with Cin7. Classifies each into an actionable issue, including
which direction a sync would go.

Read-only. Writes shopify_active_cbm_issues.csv and adds an "Active - Needs Attention"
sheet to cbm_comparison.xlsx.

Usage:
    python scripts/shopify_active_cbm_issues.py
"""

import csv
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent.parent
MAX_PLAUSIBLE_CBM = 4.0

FILLS = {
    "shop_missing": PatternFill("solid", fgColor="F4B0B0"),  # red   — Shopify has no CBM (breaks checkout)
    "cin7_missing": PatternFill("solid", fgColor="CFE2F3"),  # blue  — sync Shopify -> Cin7
    "discrepancy":  PatternFill("solid", fgColor="FFE49C"),  # amber — values disagree
    "both_missing": PatternFill("solid", fgColor="F7D6C4"),  # peach — no CBM anywhere
    "no_cin7":      PatternFill("solid", fgColor="E8E8E8"),  # grey  — no Cin7 SKU at all
}
HEADER_FILL = PatternFill("solid", fgColor="1F2A44")


def _f(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def classify_issue(status, s, c):
    """Return (issue, action, fillkey) for an active-Shopify row. s/c are floats or None."""
    s0 = s or 0
    if status == "MISSING_IN_CIN7":
        if s0 <= 0:
            return ("CBM MISSING — and no Cin7 record", "Add SKU+CBM in Cin7 (or fix SKU match)", "shop_missing")
        return ("No Cin7 record for this SKU", "Check SKU alignment / add option in Cin7", "no_cin7")
    if s0 <= 0 and (c is None or c <= 0):
        return ("CBM MISSING (blank in both)", "Enter CBM in Cin7", "both_missing")
    if s0 <= 0 and c > 0:
        tag = " — Cin7 value looks wrong (>4 m3)" if c > MAX_PLAUSIBLE_CBM else ""
        return (f"SHOPIFY CBM MISSING (Cin7 has {c}){tag}",
                "Sync Cin7 -> Shopify" if c <= MAX_PLAUSIBLE_CBM else "Fix Cin7 value first", "shop_missing")
    if c is None or c <= 0:
        return (f"CIN7 CBM MISSING (Shopify has {s0})", "Sync Shopify -> Cin7", "cin7_missing")
    # both have values
    if abs(s0 - c) <= 0.001:
        return ("", "", None)  # match — excluded
    tag = " — Cin7 value looks wrong (>4 m3)" if c > MAX_PLAUSIBLE_CBM else ""
    return (f"DISCREPANCY (Shopify {s0} vs Cin7 {c}){tag}", "Review & fix in Cin7", "discrepancy")


def main():
    rows = list(csv.DictReader((ROOT / "cbm_comparison.csv").open()))

    out = []
    for r in rows:
        status = r["status"]
        if status == "MISSING_IN_SHOPIFY":
            continue  # not an active Shopify variant
        s = _f(r["shopify_cbm"])
        c = _f(r["cin7_cbm"])
        issue, action, fillkey = classify_issue(status, s, c)
        if not issue:
            continue  # MATCH
        out.append({
            "sku": r["sku"], "product": r["product"], "option": r["option"],
            "shopify_cbm": s if s is not None else "", "cin7_cbm": c if c is not None else "",
            "issue": issue, "action": action, "_fill": fillkey,
        })

    # Order: Shopify-missing (checkout-breaking) first, then discrepancies, then Cin7-missing bulk
    order = {"shop_missing": 0, "both_missing": 1, "discrepancy": 2, "cin7_missing": 3, "no_cin7": 4}
    out.sort(key=lambda r: (order.get(r["_fill"], 9), str(r["product"])))

    # ── CSV ───────────────────────────────────────────────────────────────────
    csv_path = ROOT / "shopify_active_cbm_issues.csv"
    with csv_path.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["sku", "product", "option", "shopify_cbm", "cin7_cbm", "issue", "action"])
        w.writeheader()
        for r in out:
            w.writerow({k: r[k] for k in w.fieldnames})

    # ── add sheet to the workbook ─────────────────────────────────────────────
    xlsx_path = ROOT / "cbm_comparison.xlsx"
    wb = load_workbook(xlsx_path)
    if "Active - Needs Attention" in wb.sheetnames:
        del wb["Active - Needs Attention"]
    ws = wb.create_sheet("Active - Needs Attention", 0)  # first tab
    headers = ["SKU", "Product", "Option", "Shopify CBM", "Cin7 CBM", "Issue", "Suggested action"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    for r in out:
        ws.append([r["sku"], r["product"], r["option"], r["shopify_cbm"], r["cin7_cbm"], r["issue"], r["action"]])
        if r["_fill"]:
            for c in range(1, len(headers) + 1):
                ws.cell(ws.max_row, c).fill = FILLS[r["_fill"]]
    for i, wdt in enumerate([16, 26, 20, 12, 10, 46, 26], 1):
        ws.column_dimensions[get_column_letter(i)].width = wdt
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    wb.save(xlsx_path)

    # ── summary ───────────────────────────────────────────────────────────────
    from collections import Counter
    cnt = Counter(r["_fill"] for r in out)
    print(f"{len(out)} active Shopify variants need attention -> {csv_path.name} + xlsx tab")
    labels = {"shop_missing": "Shopify CBM MISSING", "both_missing": "Missing in both",
              "discrepancy": "Discrepancy (both have, differ)", "cin7_missing": "Cin7 CBM MISSING (Shopify has it)",
              "no_cin7": "No Cin7 record"}
    for k in ["shop_missing", "both_missing", "discrepancy", "cin7_missing", "no_cin7"]:
        print(f"  {labels[k]:36} {cnt.get(k, 0)}")


if __name__ == "__main__":
    main()
